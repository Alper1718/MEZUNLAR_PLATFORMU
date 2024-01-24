import tkinter as tk
from tkinter import messagebox
from PIL import Image, ImageTk
import openpyxl
from openpyxl import Workbook
import os

class StudentManager:
    def __init__(self, filename='mezunlar.xlsx'):
        self.filename = filename
        self.workbook = self.create_workbook_if_not_exists()

    def create_workbook_if_not_exists(self):
        try:
            workbook = openpyxl.load_workbook(self.filename)
        except FileNotFoundError:
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["İsim", "Üniversite", "Mezun Olma Yılı", "Bölüm", "Telefon", "Mail", "Adres", "Çalıştığı Kurum"])
            workbook.save(self.filename)
        return workbook

    def add_student(self, name, university, year, faculty, phone, email, address, workplace):
        sheet = self.workbook.active
        sheet.append([name, university, year, faculty, phone, email, address, workplace])
        self.workbook.save(self.filename)

    def search_student(self, search_name):
        sheet = self.workbook.active
        found = False
        result = ""

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] == search_name:
                result += f"İsim: {row[0]}, Kazandığı Üniversite: {row[1]}, Mezun Olma Yılı: {row[2]}, Kazandığı Bölüm: {row[3]}, "
                result += f"Telefon: {row[4]}, Mail: {row[5]}, Adres: {row[6]}, Çalıştığı Kurum: {row[7]}\n"
                found = True

        if not found:
            result = "Öğrenci bulunamadı."

        return result

    def remove_student(self, remove_name):
        sheet = self.workbook.active
        found = False
        result = ""

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] == remove_name:
                result += f"İsim: {row[0]}, Kazandığı Üniversite: {row[1]}, Mezun olma yılı: {row[2]}, Kazandığı bölüm: {row[3]}, "
                result += f"Telefon: {row[4]}, Mail: {row[5]}, Adres: {row[6]}, Çalıştığı Kurum: {row[7]}\n"
                found = True

        if not found:
            result = "Öğrenci bulunamadı."
        else:
            confirm = messagebox.askyesno("Onay", f"Bu öğrenci silmek istediğinizden emin misiniz?\n{result}")

            if confirm:
                for row in sheet.iter_rows(min_row=2):
                    if row[0].value == remove_name:
                        sheet.delete_rows(row[0].row)
                        self.workbook.save(self.filename)
                        messagebox.showinfo("Başarılı", "Mezun başarıyla kaldırıldı.")
                        break


class StudentManagementGUI:
    def __init__(self, root, student_manager, image_left_path, image_right_path):
        self.root = root
        self.student_manager = student_manager
        self.image_left_path = image_left_path
        self.image_right_path = image_right_path
        self.initialize_gui()

    def initialize_gui(self):
        self.root.title("AMAL Mezun Yönetim Sistemi")

        # Load images
        img_left = Image.open(self.image_left_path)
        img_right = Image.open(self.image_right_path)

        # Resize images if needed
        img_left = img_left.resize((450, 300))
        img_right = img_right.resize((350, 350))

        # Convert images to Tkinter PhotoImage objects
        img_left_tk = ImageTk.PhotoImage(img_left)
        img_right_tk = ImageTk.PhotoImage(img_right)

        # Create labels for images and place them in the top left and top right corners
        label_left = tk.Label(self.root, image=img_left_tk)
        label_left.photo = img_left_tk
        label_left.pack(side=tk.LEFT, padx=5, pady=5)

        label_right = tk.Label(self.root, image=img_right_tk)
        label_right.photo = img_right_tk
        label_right.pack(side=tk.RIGHT, padx=5, pady=5)

        self.label = tk.Label(self.root, text="Seçenekler:", font=("Helvetica", 16))
        self.label.pack()

        button_height = 2
        button_width = 20

        self.add_button = tk.Button(self.root, text="Mezun Ekle", command=self.add_student, height=button_height, width=button_width, font=("Helvetica", 12))
        self.add_button.pack()

        self.search_button = tk.Button(self.root, text="Mezun Ara", command=self.search_student, height=button_height, width=button_width, font=("Helvetica", 12))
        self.search_button.pack()

        self.remove_button = tk.Button(self.root, text="Mezun Sil", command=self.remove_student, height=button_height, width=button_width, font=("Helvetica", 12))
        self.remove_button.pack()

        self.exit_button = tk.Button(self.root, text="Çıkış", command=self.root.destroy, height=button_height, width=button_width, font=("Helvetica", 12))
        self.exit_button.pack()

    def add_student(self):
        self.clear_frame()

        tk.Label(self.root, text="İsim:", font=("Helvetica", 12)).pack()
        name_entry = tk.Entry(self.root)
        name_entry.pack()

        tk.Label(self.root, text="Kazandığı üniversite:", font=("Helvetica", 12)).pack()
        university_entry = tk.Entry(self.root)
        university_entry.pack()

        tk.Label(self.root, text="Mezun olma yılı:", font=("Helvetica", 12)).pack()
        year_entry = tk.Entry(self.root)
        year_entry.pack()

        tk.Label(self.root, text="Fakülte:", font=("Helvetica", 12)).pack()
        faculty_entry = tk.Entry(self.root)
        faculty_entry.pack()

        tk.Label(self.root, text="Telefon:", font=("Helvetica", 12)).pack()
        phone_entry = tk.Entry(self.root)
        phone_entry.pack()

        tk.Label(self.root, text="Mail:", font=("Helvetica", 12)).pack()
        email_entry = tk.Entry(self.root)
        email_entry.pack()

        tk.Label(self.root, text="Adres:", font=("Helvetica", 12)).pack()
        address_entry = tk.Entry(self.root)
        address_entry.pack()

        tk.Label(self.root, text="Çalıştığı Kurum:", font=("Helvetica", 12)).pack()
        workplace_entry = tk.Entry(self.root)
        workplace_entry.pack()

        save_button = tk.Button(self.root, text="Kaydet",
                                command=lambda: self.save_student(name_entry, university_entry, year_entry,
                                                                  faculty_entry, phone_entry, email_entry, address_entry, workplace_entry),
                                font=("Helvetica", 12))
        save_button.pack()

    def save_student(self, name_entry, university_entry, year_entry, faculty_entry, phone_entry, email_entry, address_entry, workplace_entry):
        name = name_entry.get()
        university = university_entry.get()
        year = year_entry.get()
        faculty = faculty_entry.get()
        phone = phone_entry.get()
        email = email_entry.get()
        address = address_entry.get()
        workplace = workplace_entry.get()

        if not all([name, university, year, faculty, phone, email, address, workplace]):
            messagebox.showerror("Hata", "Tüm alanlar dolu olmalı.")
            return

        try:
            year = int(year)
        except ValueError:
            messagebox.showerror("Hata", "Mezun olma yılı bir sayı olmalıdır.")
            return

        self.student_manager.add_student(name, university, year, faculty, phone, email, address, workplace)
        messagebox.showinfo("Başarılı", "Mezun başarıyla eklendi.")

        self.clear_frame()
        self.initialize_gui()

    def search_student(self):
        self.clear_frame()

        tk.Label(self.root, text="Mezun aramak için isim girin", font=("Helvetica", 12)).pack()
        search_entry = tk.Entry(self.root)
        search_entry.pack()

        search_button = tk.Button(self.root, text="Ara", command=lambda: self.display_search_result(search_entry), font=("Helvetica", 12))
        search_button.pack()

    def display_search_result(self, search_entry):
        search_name = search_entry.get()
        result = self.student_manager.search_student(search_name)
        messagebox.showinfo("Arama sonucu:", result)

        self.clear_frame()
        self.initialize_gui()

    def remove_student(self):
        self.clear_frame()

        tk.Label(self.root, text="Mezun kaldırmak için isim girin:", font=("Helvetica", 12)).pack()
        remove_entry = tk.Entry(self.root)
        remove_entry.pack()

        remove_button = tk.Button(self.root, text="Kaldır", command=lambda: self.confirm_remove(remove_entry), font=("Helvetica", 12))
        remove_button.pack()

    def confirm_remove(self, remove_entry):
        remove_name = remove_entry.get()
        self.student_manager.remove_student(remove_name)
        remove_entry.delete(0, tk.END)

        self.clear_frame()
        self.initialize_gui()

    def clear_frame(self):
        for widget in self.root.winfo_children():
            widget.destroy()

def main():
    root = tk.Tk()

    # Get the script's directory
    script_dir = os.path.dirname(os.path.abspath(__file__))

    # Construct image paths relative to the script's directory
    image_left_path = os.path.join(script_dir, "hawks.png")
    image_right_path = os.path.join(script_dir, "amal.png")

    student_manager = StudentManager()
    app = StudentManagementGUI(root, student_manager, image_left_path, image_right_path)
    root.mainloop()

if __name__ == "__main__":
    main()
