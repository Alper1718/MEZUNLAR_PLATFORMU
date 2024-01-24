import tkinter as tk
from tkinter import messagebox
import openpyxl
from openpyxl import Workbook

class StudentManager:
    def __init__(self, filename='students.xlsx'):
        self.filename = filename
        self.workbook = self.create_workbook_if_not_exists()

    def create_workbook_if_not_exists(self):
        try:
            workbook = openpyxl.load_workbook(self.filename)
        except FileNotFoundError:
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["Name", "University", "Year", "Faculty"])
            workbook.save(self.filename)
        return workbook

    def add_student(self, name, university, year, faculty):
        sheet = self.workbook.active
        sheet.append([name, university, year, faculty])
        self.workbook.save(self.filename)

    def search_student(self, search_name):
        sheet = self.workbook.active
        found = False
        result = ""

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] == search_name:
                result += f"Name: {row[0]}, University: {row[1]}, Year: {row[2]}, Faculty: {row[3]}\n"
                found = True

        if not found:
            result = "Student not found."

        return result

    def remove_student(self, remove_name):
        sheet = self.workbook.active
        found = False
        result = ""

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] == remove_name:
                result += f"Name: {row[0]}, University: {row[1]}, Year: {row[2]}, Faculty: {row[3]}\n"
                found = True

        if not found:
            result = "Student not found."
        else:
            confirm = messagebox.askyesno("Confirmation", f"Are you sure you want to delete this student?\n{result}")

            if confirm:
                for row in sheet.iter_rows(min_row=2):
                    if row[0].value == remove_name:
                        sheet.delete_rows(row[0].row)
                        self.workbook.save(self.filename)
                        messagebox.showinfo("Success", "Student removed successfully.")
                        break

class StudentManagementGUI:
    def __init__(self, root, student_manager):
        self.root = root
        self.student_manager = student_manager
        self.initialize_gui()

    def initialize_gui(self):
        self.root.title("Student Management System")

        self.label = tk.Label(self.root, text="Options:")
        self.label.pack()

        self.add_button = tk.Button(self.root, text="Add Student", command=self.add_student, height=2)
        self.add_button.pack()

        self.search_button = tk.Button(self.root, text="Search Student", command=self.search_student, height=2)
        self.search_button.pack()

        self.remove_button = tk.Button(self.root, text="Remove Student", command=self.remove_student, height=2)
        self.remove_button.pack()

        self.exit_button = tk.Button(self.root, text="Exit", command=self.root.destroy, height=2)
        self.exit_button.pack()

    def add_student(self):
        self.clear_frame()
        
        tk.Label(self.root, text="Name:").pack()
        name_entry = tk.Entry(self.root)
        name_entry.pack()

        tk.Label(self.root, text="University:").pack()
        university_entry = tk.Entry(self.root)
        university_entry.pack()

        tk.Label(self.root, text="Year of Graduation:").pack()
        year_entry = tk.Entry(self.root)
        year_entry.pack()

        tk.Label(self.root, text="Faculty:").pack()
        faculty_entry = tk.Entry(self.root)
        faculty_entry.pack()

        save_button = tk.Button(self.root, text="Save", command=lambda: self.save_student(name_entry, university_entry, year_entry, faculty_entry))
        save_button.pack()

    def save_student(self, name_entry, university_entry, year_entry, faculty_entry):
        name = name_entry.get()
        university = university_entry.get()
        year = year_entry.get()
        faculty = faculty_entry.get()

        if not all([name, university, year, faculty]):
            messagebox.showerror("Error", "All fields must be filled.")
            return

        try:
            year = int(year)
        except ValueError:
            messagebox.showerror("Error", "Year of Graduation must be a valid integer.")
            return

        self.student_manager.add_student(name, university, year, faculty)
        messagebox.showinfo("Success", "Student added successfully.")

        self.clear_frame()
        self.initialize_gui()

    def search_student(self):
        self.clear_frame()

        tk.Label(self.root, text="Enter the name to search:").pack()
        search_entry = tk.Entry(self.root)
        search_entry.pack()

        search_button = tk.Button(self.root, text="Search", command=lambda: self.display_search_result(search_entry))
        search_button.pack()

    def display_search_result(self, search_entry):
        search_name = search_entry.get()
        result = self.student_manager.search_student(search_name)
        messagebox.showinfo("Search Result", result)

        self.clear_frame()
        self.initialize_gui()

    def remove_student(self):
        self.clear_frame()

        tk.Label(self.root, text="Enter the name to remove:").pack()
        remove_entry = tk.Entry(self.root)
        remove_entry.pack()

        remove_button = tk.Button(self.root, text="Remove", command=lambda: self.confirm_remove(remove_entry))
        remove_button.pack()

    def confirm_remove(self, remove_entry):
        remove_name = remove_entry.get()
        self.student_manager.remove_student(remove_name)
        remove_entry.delete(0, tk.END)

        self.clear_frame()
        self.initialize_gui()

    def clear_frame(self):
        for widget in self.root.winfo_children():
            widget.destroy()

def main():
    root = tk.Tk()
    student_manager = StudentManager()
    app = StudentManagementGUI(root, student_manager)
    root.mainloop()

if __name__ == "__main__":
    main()
