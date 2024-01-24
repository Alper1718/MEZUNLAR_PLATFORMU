import tkinter as tk
from tkinter import messagebox
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

class StudentManagementGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("Student Management System")

        self.create_workbook_if_not_exists()

        self.label = tk.Label(root, text="Options:")
        self.label.pack()

        self.add_button = tk.Button(root, text="Add Student", command=self.add_student)
        self.add_button.pack()

        self.search_button = tk.Button(root, text="Search Student", command=self.search_student)
        self.search_button.pack()

        self.remove_button = tk.Button(root, text="Remove Student", command=self.remove_student)
        self.remove_button.pack()

        self.exit_button = tk.Button(root, text="Exit", command=root.destroy)
        self.exit_button.pack()

    def create_workbook_if_not_exists(self):
        try:
            workbook = openpyxl.load_workbook('students.xlsx')
        except FileNotFoundError:
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["Name", "University", "Year", "Faculty"])
            workbook.save('students.xlsx')

    def add_student(self):
        add_window = tk.Toplevel(self.root)
        add_window.title("Add Student")

        name_label = tk.Label(add_window, text="Name:")
        name_label.grid(row=0, column=0)
        self.name_entry = tk.Entry(add_window)
        self.name_entry.grid(row=0, column=1)

        university_label = tk.Label(add_window, text="University:")
        university_label.grid(row=1, column=0)
        self.university_entry = tk.Entry(add_window)
        self.university_entry.grid(row=1, column=1)

        year_label = tk.Label(add_window, text="Year of Graduation:")
        year_label.grid(row=2, column=0)
        self.year_entry = tk.Entry(add_window)
        self.year_entry.grid(row=2, column=1)

        faculty_label = tk.Label(add_window, text="Faculty:")
        faculty_label.grid(row=3, column=0)
        self.faculty_entry = tk.Entry(add_window)
        self.faculty_entry.grid(row=3, column=1)

        save_button = tk.Button(add_window, text="Save", command=self.save_student)
        save_button.grid(row=4, columnspan=2)

    def save_student(self):
        name = self.name_entry.get()
        university = self.university_entry.get()
        year = self.year_entry.get()
        faculty = self.faculty_entry.get()

        workbook = openpyxl.load_workbook('students.xlsx')
        sheet = workbook.active
        sheet.append([name, university, year, faculty])
        workbook.save('students.xlsx')
        messagebox.showinfo("Success", "Student added successfully.")
        self.name_entry.delete(0, tk.END)
        self.university_entry.delete(0, tk.END)
        self.year_entry.delete(0, tk.END)
        self.faculty_entry.delete(0, tk.END)

    def search_student(self):
        search_window = tk.Toplevel(self.root)
        search_window.title("Search Student")

        search_label = tk.Label(search_window, text="Enter the name to search:")
        search_label.pack()

        self.search_entry = tk.Entry(search_window)
        self.search_entry.pack()

        search_button = tk.Button(search_window, text="Search", command=self.display_search_result)
        search_button.pack()

    def display_search_result(self):
        search_name = self.search_entry.get()

        workbook = openpyxl.load_workbook('students.xlsx')
        sheet = workbook.active

        found = False
        result = ""

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] == search_name:
                result += f"Name: {row[0]}, University: {row[1]}, Year: {row[2]}, Faculty: {row[3]}\n"
                found = True

        if not found:
            result = "Student not found."

        messagebox.showinfo("Search Result", result)

    def remove_student(self):
        remove_window = tk.Toplevel(self.root)
        remove_window.title("Remove Student")

        remove_label = tk.Label(remove_window, text="Enter the name to remove:")
        remove_label.pack()

        self.remove_entry = tk.Entry(remove_window)
        self.remove_entry.pack()

        remove_button = tk.Button(remove_window, text="Remove", command=self.confirm_remove)
        remove_button.pack()

    def confirm_remove(self):
        remove_name = self.remove_entry.get()

        workbook = openpyxl.load_workbook('students.xlsx')
        sheet = workbook.active

        found = False
        result = ""

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] == remove_name:
                result += f"Name: {row[0]}, University: {row[1]}, Year: {row[2]}, Faculty: {row[3]}\n"
                found = True

        if not found:
            result = "Student not found."
        else:
            confirm = messagebox.askyesno("Confirmation", f"Are you sure you want to delete this student?\n{result}")

            if confirm:
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    if row[0] == remove_name:
                        sheet.delete_rows(row[0].row)
                        workbook.save('students.xlsx')
                        messagebox.showinfo("Success", "Student removed successfully.")
                        break

        self.remove_entry.delete(0, tk.END)

def main():
    root = tk.Tk()
    app = StudentManagementGUI(root)
    root.mainloop()

if __name__ == "__main__":
    main()
