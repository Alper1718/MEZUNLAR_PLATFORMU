import tkinter as tk
from tkinter import messagebox
from PIL import Image, ImageTk
from openpyxl import Workbook, load_workbook
import os
import hashlib
import time


def check_license():
    """Check if the hash of the image matches the expected hash."""
    expected_hash = "8464b97a81e4e221f9b99c5639488e7e"  # Replace this with the expected hash value
    relative_path = "mezun programı/alişahin/licence.png"  # Replace this with the relative path to your license image
    script_dir = os.path.dirname(os.path.realpath(__file__))
    image_path = os.path.join(script_dir, relative_path)

    if os.path.exists(image_path):
        image_hash = StudentManager.calculate_hash(image_path)
        if image_hash == expected_hash:
            print("Erişim izni verildi.")
            return True
        else:
            print("Lisans geçerli değil.")
            return False
    else:
        print("Lisans bulunamadı.")
        return False


class StudentManager:
    def __init__(self, filename='mezunlar.xlsx'):
        self.filename = filename
        self.workbook = self.create_workbook_if_not_exists()

    def create_workbook_if_not_exists(self):
        try:
            workbook = load_workbook(self.filename)
        except FileNotFoundError:
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["İsim", "Üniversite", "Mezun Olma Yılı", "Bölüm", "Telefon", "Mail", "Adres", "Çalıştığı Kurum"])
            workbook.save(self.filename)
        return workbook

    @staticmethod
    def calculate_hash(file_path):
        """Calculate the hash of an image file."""
        with open(file_path, 'rb') as f:
            image_data = f.read()
            return hashlib.md5(image_data).hexdigest()

    def add_student(self, name, university, year, faculty, phone, email, address, workplace):
        sheet = self.workbook.active
        sheet.append([name, university, year, faculty, phone, email, address, workplace])
        self.workbook.save(self.filename)

    def search_student(self, search_name):
        found = False
        result = ""

        for sheet_name in self.workbook.sheetnames:
            sheet = self.workbook[sheet_name]
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row[0] == search_name:
                    result += f"İsim: {row[0]}, Kazandığı Üniversite: {row[1]}, Mezun Olma Yılı: {row[2]}, Kazandığı Bölüm: {row[3]}, "
                    result += f"Telefon: {row[4]}, Mail: {row[5]}, Adres: {row[6]}, Çalıştığı Kurum: {row[7]}\n"
                    found = True

        if not found:
            result = "Öğrenci bulunamadı."

        return result

    def remove_student(self, remove_name):
        sheet = self.workbook.active
        found = False
        result = ""

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] == remove_name:
                result += f"İsim: {row[0]}, Kazandığı Üniversite: {row[1]}, Mezun olma yılı: {row[2]}, Kazandığı bölüm: {row[3]}, "
                result += f"Telefon: {row[4]}, Mail: {row[5]}, Adres: {row[6]}, Çalıştığı Kurum: {row[7]}\n"
                found = True

        if not found:
            result = "Öğrenci bulunamadı."
        else:
            confirm = messagebox.askyesno("Onay", f"Bu öğrenci silmek istediğinizden emin misiniz?\n{result}")

            if confirm:
                for row in sheet.iter_rows(min_row=2):
                    if row[0].value == remove_name:
                        sheet.delete_rows(row[0].row)
                        self.workbook.save(self.filename)
                        messagebox.showinfo("Başarılı", "Mezun başarıyla kaldırıldı.")
                        break

    def filter_by_year(self, filter_year):
        sheet = self.workbook.active
        filtered_students = []

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[2] == filter_year:
                filtered_students.append(row)

        return filtered_students

    def export_filtered_students(self, filtered_students, new_filename):
        new_workbook = Workbook()
        new_sheet = new_workbook.active
        new_sheet.append(["İsim", "Üniversite", "Mezun Olma Yılı", "Bölüm", "Telefon", "Mail", "Adres", "Çalıştığı Kurum"])

        for student in filtered_students:
            new_sheet.append(student)

        new_workbook.save(new_filename)


class StudentManagementGUI:
    def __init__(self, root, student_manager, image_left_path, image_right_path):
        self.root = root
        self.student_manager = student_manager
        self.image_left_path = image_left_path
        self.image_right_path = image_right_path
        self.img_left = None
        self.img_right = None
        self.initialize_gui()

    def load_images(self):
        img_left = Image.open(self.image_left_path)
        img_right = Image.open(self.image_right_path)
        self.img_left = ImageTk.PhotoImage(img_left.resize((330, 300)))
        self.img_right = ImageTk.PhotoImage(img_right.resize((325, 325)))

    def initialize_gui(self):
        self.clear_frame()

        self.root.title("Adnan Menderes Anadolu Lisesi Mezun Yönetim Sistemi")
        self.root.attributes('-fullscreen', True)
        self.load_images()

        label_left = tk.Label(self.root, image=self.img_left)
        label_left.pack(side="left", anchor="s")

        label_right = tk.Label(self.root, image=self.img_right)
        label_right.pack(side="right", anchor="s")

        self.baslik = tk.Label(self.root, text="Adnan Menderes Anadolu Lisesi Mezun Yönetim Sistemi",
                               font=("Helvetica", 25), pady=20)
        self.baslik.pack()
        self.label = tk.Label(self.root, text="Seçenekler:", font=("Helvetica", 16))
        self.label.pack()

        button_height = 50
        button_width = 250

        # Load icons
        icon_add = Image.open("singleplayer.png").resize((40, 40))
        icon_search = Image.open("zoom.png").resize((40, 40))
        icon_remove = Image.open("remove.png").resize((40, 40))
        icon_filter = Image.open("export.png").resize((40, 40))
        icon_exit = Image.open("exit.png").resize((40, 40))

        self.icon_add = ImageTk.PhotoImage(icon_add)
        self.icon_search = ImageTk.PhotoImage(icon_search)
        self.icon_remove = ImageTk.PhotoImage(icon_remove)
        self.icon_filter = ImageTk.PhotoImage(icon_filter)
        self.icon_exit = ImageTk.PhotoImage(icon_exit)

        # Add Mezun Ekle button
        self.add_button = tk.Button(self.root, text="Mezun Ekle", command=self.add_student, font=("Helvetica", 12),
                                    height=button_height, width=button_width)
        self.add_button.config(image=self.icon_add, compound=tk.LEFT)
        self.add_button.pack(anchor="center")

        # Add Mezun Ara button
        self.search_button = tk.Button(self.root, text="Mezun Ara", command=self.search_student, font=("Helvetica", 12),
                                       height=button_height, width=button_width)
        self.search_button.config(image=self.icon_search, compound=tk.LEFT)
        self.search_button.pack(anchor="center")

        # Add Mezun Sil button
        self.remove_button = tk.Button(self.root, text="Mezun Sil", command=self.remove_student, font=("Helvetica", 12),
                                       height=button_height, width=button_width)
        self.remove_button.config(image=self.icon_remove, compound=tk.LEFT)
        self.remove_button.pack(anchor="center")

        # Add Filtrele ve Dosya Oluştur button
        self.filter_button = tk.Button(self.root, text="Filtrele ve Dosya Oluştur", command=self.filter_by_year,
                                       font=("Helvetica", 12), height=button_height, width=button_width)
        self.filter_button.config(image=self.icon_filter, compound=tk.LEFT)
        self.filter_button.pack(anchor="center")

        # Add Çıkış button
        self.exit_button = tk.Button(self.root, text="Çıkış", command=self.root.destroy, font=("Helvetica", 12),
                                     height=button_height, width=button_width)
        self.exit_button.config(image=self.icon_exit, compound=tk.LEFT)
        self.exit_button.pack(anchor="center")

    def add_student(self):

        button_height = 50
        button_width = 250

        self.clear_frame()
        tk.Label(self.root, text="İsim:", font=("Helvetica", 14)).pack()
        name_entry = tk.Entry(self.root)
        name_entry.pack()

        tk.Label(self.root, text="Kazandığı üniversite:", font=("Helvetica", 14)).pack()
        university_entry = tk.Entry(self.root)
        university_entry.pack()

        tk.Label(self.root, text="Mezun olma yılı:", font=("Helvetica", 14)).pack()
        year_entry = tk.Entry(self.root)
        year_entry.pack()

        tk.Label(self.root, text="Fakülte:", font=("Helvetica", 14)).pack()
        faculty_entry = tk.Entry(self.root)
        faculty_entry.pack()

        tk.Label(self.root, text="Telefon:", font=("Helvetica", 14)).pack()
        phone_entry = tk.Entry(self.root)
        phone_entry.pack()

        tk.Label(self.root, text="Mail:", font=("Helvetica", 14)).pack()
        email_entry = tk.Entry(self.root)
        email_entry.pack()

        tk.Label(self.root, text="Adres:", font=("Helvetica", 14)).pack()
        address_entry = tk.Entry(self.root)
        address_entry.pack()

        tk.Label(self.root, text="Çalıştığı Kurum:", font=("Helvetica", 14)).pack()
        workplace_entry = tk.Entry(self.root)
        workplace_entry.pack()

        tk.Label(self.root, text="").pack()

        save_button = tk.Button(self.root, text="Kaydet        ",
                                command=lambda: self.save_student(name_entry, university_entry, year_entry,
                                                                  faculty_entry, phone_entry, email_entry,
                                                                  address_entry, workplace_entry),
                                font=("Helvetica", 12), height=button_height, width=button_width)

        # Load the image for the save button
        icon_add = Image.open("singleplayer.png").resize((40, 40))
        self.icon_add = ImageTk.PhotoImage(icon_add)
        save_button.config(image=self.icon_add, compound=tk.LEFT)  # Configure the button's image
        save_button.pack()

        cancel_button = tk.Button(self.root, text="İptal        ", command=self.initialize_gui, font=("Helvetica", 12),
                                  height=button_height, width=button_width)

        # Load the image for the cancel button
        icon_cancel = Image.open("cross.png").resize((40, 40))
        self.icon_cancel = ImageTk.PhotoImage(icon_cancel)
        cancel_button.config(image=self.icon_cancel, compound=tk.LEFT)  # Configure the button's image
        cancel_button.pack()

        img_left = Image.open(self.image_left_path)
        img_right = Image.open(self.image_right_path)

        img_left = img_left.resize((330, 300))
        img_right = img_right.resize((325, 325))

        img_left_tk = ImageTk.PhotoImage(img_left)
        img_right_tk = ImageTk.PhotoImage(img_right)

        label_left = tk.Label(self.root, image=img_left_tk)
        label_left.photo = img_left_tk
        label_left.pack(side="left", anchor="s")

        label_right = tk.Label(self.root, image=img_right_tk)
        label_right.photo = img_right_tk
        label_right.pack(side="right", anchor="s")

    def save_student(self, name_entry, university_entry, year_entry, faculty_entry, phone_entry, email_entry, address_entry, workplace_entry):
        name = name_entry.get()
        university = university_entry.get()
        year = year_entry.get()
        faculty = faculty_entry.get()
        phone = phone_entry.get()
        email = email_entry.get()
        address = address_entry.get()
        workplace = workplace_entry.get()

        if not all([name, university, year, faculty, phone, email, address, workplace]):
            messagebox.showerror("Hata", "Tüm alanlar dolu olmalı.")
            return

        try:
            year = int(year)
        except ValueError:
            messagebox.showerror("Hata", "Mezun olma yılı bir sayı olmalıdır.")
            return

        self.student_manager.add_student(name, university, year, faculty, phone, email, address, workplace)
        messagebox.showinfo("Başarılı", "Mezun başarıyla eklendi.")

        self.clear_frame()
        self.initialize_gui()

    def search_student(self):
        self.clear_frame()

        button_height = 50
        button_width = 250

        tk.Label(self.root, text="Mezun aramak için isim girin", font=("Helvetica", 16), pady=20).pack()
        search_entry = tk.Entry(self.root)
        search_entry.pack()

        tk.Label(self.root, text="").pack()

        search_button = tk.Button(self.root, text="Ara        ", command=lambda: self.display_search_result(search_entry),
                                  font=("Helvetica", 12), height=button_height, width=button_width)

        # Load the image for the search button
        icon_search = Image.open("zoom.png").resize((40, 40))
        self.icon_search = ImageTk.PhotoImage(icon_search)
        search_button.config(image=self.icon_search, compound=tk.LEFT)  # Configure the button's image
        search_button.pack()

        cancel_button = tk.Button(self.root, text="İptal        ", command=self.initialize_gui, font=("Helvetica", 12),
                                  height=button_height, width=button_width)

        # Load the image for the cancel button
        icon_cancel = Image.open("cross.png").resize((40, 40))
        self.icon_cancel = ImageTk.PhotoImage(icon_cancel)
        cancel_button.config(image=self.icon_cancel, compound=tk.LEFT)  # Configure the button's image
        cancel_button.pack()

        img_left = Image.open(self.image_left_path)
        img_right = Image.open(self.image_right_path)

        img_left = img_left.resize((330, 300))
        img_right = img_right.resize((325, 325))

        img_left_tk = ImageTk.PhotoImage(img_left)
        img_right_tk = ImageTk.PhotoImage(img_right)

        label_left = tk.Label(self.root, image=img_left_tk)
        label_left.photo = img_left_tk
        label_left.pack(side="left", anchor="s")

        label_right = tk.Label(self.root, image=img_right_tk)
        label_right.photo = img_right_tk
        label_right.pack(side="right", anchor="s")

    def display_search_result(self, search_entry):
        search_name = search_entry.get()
        result = self.student_manager.search_student(search_name)
        messagebox.showinfo("Arama sonucu:", result)

        self.clear_frame()
        self.initialize_gui()

    def remove_student(self):
        self.clear_frame()

        button_height = 50
        button_width = 250

        tk.Label(self.root, text="Mezun kaldırmak için isim girin:", font=("Helvetica", 14)).pack()
        remove_entry = tk.Entry(self.root)
        remove_entry.pack()

        tk.Label(self.root, text="").pack()

        remove_button = tk.Button(self.root, text="Kaldır      ", command=lambda: self.confirm_remove(remove_entry),
                                  font=("Helvetica", 12), height=button_height, width=button_width)

        # Load the image for the remove button
        icon_remove = Image.open("trashcan.png").resize((40, 40))
        self.icon_remove = ImageTk.PhotoImage(icon_remove)
        remove_button.config(image=self.icon_remove, compound=tk.LEFT)  # Configure the button's image
        remove_button.pack()

        cancel_button = tk.Button(self.root, text="İptal      ", command=self.initialize_gui, font=("Helvetica", 12),
                                  height=button_height, width=button_width)

        # Load the image for the cancel button
        icon_cancel = Image.open("cross.png").resize((40, 40))
        self.icon_cancel = ImageTk.PhotoImage(icon_cancel)
        cancel_button.config(image=self.icon_cancel, compound=tk.LEFT)  # Configure the button's image
        cancel_button.pack()

        img_left = Image.open(self.image_left_path)
        img_right = Image.open(self.image_right_path)

        img_left = img_left.resize((330, 300))
        img_right = img_right.resize((325, 325))

        img_left_tk = ImageTk.PhotoImage(img_left)
        img_right_tk = ImageTk.PhotoImage(img_right)

        label_left = tk.Label(self.root, image=img_left_tk)
        label_left.photo = img_left_tk
        label_left.pack(side="left", anchor="s")

        label_right = tk.Label(self.root, image=img_right_tk)
        label_right.photo = img_right_tk
        label_right.pack(side="right", anchor="s")

    def confirm_remove(self, remove_entry):
        remove_name = remove_entry.get()
        self.student_manager.remove_student(remove_name)
        remove_entry.delete(0, tk.END)

        self.clear_frame()
        self.initialize_gui()

    def filter_by_year(self):
        self.clear_frame()

        self.filter_label = tk.Label(self.root, text="Filtreleme Türünü Seçin:", font=("Helvetica", 16), pady=20)
        self.filter_label.pack()

        button_height = 50
        button_width = 250

        # Load images for each button
        icon_filter_year = Image.open("button1.png").resize((40, 40))
        icon_filter_university = Image.open("button2.png").resize((40, 40))
        icon_filter_faculty = Image.open("button3.png").resize((40, 40))
        icon_exit = Image.open("cross.png").resize((40, 40))

        self.icon_filter_year = ImageTk.PhotoImage(icon_filter_year)
        self.icon_filter_university = ImageTk.PhotoImage(icon_filter_university)
        self.icon_filter_faculty = ImageTk.PhotoImage(icon_filter_faculty)
        self.icon_exit = ImageTk.PhotoImage(icon_exit)

        filter_year_button = tk.Button(self.root, text="Yıla Göre Filtrele", command=lambda: self.prompt_filter("yıl"),
                                       height=button_height, width=button_width, font=("Helvetica", 12))
        filter_year_button.config(image=self.icon_filter_year, compound=tk.LEFT)  # Configure the button's image
        filter_year_button.pack()

        filter_university_button = tk.Button(self.root, text="Üniversiteye Göre Filtrele",
                                             command=lambda: self.prompt_filter("universite"), height=button_height,
                                             width=button_width, font=("Helvetica", 12))
        filter_university_button.config(image=self.icon_filter_university,
                                        compound=tk.LEFT)  # Configure the button's image
        filter_university_button.pack()

        filter_faculty_button = tk.Button(self.root, text="Fakülteye Göre Filtrele",
                                          command=lambda: self.prompt_filter("fakulte"), height=button_height,
                                          width=button_width, font=("Helvetica", 12))
        filter_faculty_button.config(image=self.icon_filter_faculty, compound=tk.LEFT)  # Configure the button's image
        filter_faculty_button.pack()

        exit_button = tk.Button(self.root, text="İptal", command=self.initialize_gui, height=button_height,
                                width=button_width, font=("Helvetica", 12))
        exit_button.config(image=self.icon_exit, compound=tk.LEFT)  # Configure the button's image
        exit_button.pack()

        img_left = Image.open(self.image_left_path)
        img_right = Image.open(self.image_right_path)

        img_left = img_left.resize((330, 300))
        img_right = img_right.resize((325, 325))

        img_left_tk = ImageTk.PhotoImage(img_left)
        img_right_tk = ImageTk.PhotoImage(img_right)

        label_left = tk.Label(self.root, image=img_left_tk)
        label_left.photo = img_left_tk
        label_left.pack(side="left", anchor="s")

        label_right = tk.Label(self.root, image=img_right_tk)
        label_right.photo = img_right_tk
        label_right.pack(side="right", anchor="s")

    def prompt_filter(self, filter_type):
        self.clear_frame()

        button_height = 50
        button_width = 250

        filter_label_text = ""

        if filter_type == "yıl":
            filter_label_text = "Filtrelemelek istediğiniz yılı girin"
        elif filter_type == "universite":
            filter_label_text = "Filtrelemelek istediğiniz üniversiteyi girin"
        elif filter_type == "fakulte":
            filter_label_text = "Filtrelemelek istediğiniz fakülteyi girin"

        tk.Label(self.root, text=filter_label_text, font=("Helvetica", 14)).pack()
        filter_entry = tk.Entry(self.root)
        filter_entry.pack()

        tk.Label(self.root, text="").pack()

        # Load images for the buttons
        icon_filter = Image.open("filter-removebg-preview.png").resize((40, 40))
        icon_cancel = Image.open("cross.png").resize((40, 40))

        self.icon_filter = ImageTk.PhotoImage(icon_filter)
        self.icon_cancel = ImageTk.PhotoImage(icon_cancel)

        filter_button = tk.Button(self.root, text="Filtrele",
                                  command=lambda: self.apply_filter(filter_type, filter_entry), font=("Helvetica", 12),
                                  height=button_height, width=button_width)
        filter_button.config(image=self.icon_filter, compound=tk.LEFT)  # Configure the button's image
        filter_button.pack()

        cancel_button = tk.Button(self.root, text="İptal", command=self.initialize_gui, font=("Helvetica", 12),
                                  height=button_height, width=button_width)
        cancel_button.config(image=self.icon_cancel, compound=tk.LEFT)  # Configure the button's image
        cancel_button.pack()

        img_left = Image.open(self.image_left_path)
        img_right = Image.open(self.image_right_path)

        img_left = img_left.resize((330, 300))
        img_right = img_right.resize((325, 325))

        img_left_tk = ImageTk.PhotoImage(img_left)
        img_right_tk = ImageTk.PhotoImage(img_right)

        label_left = tk.Label(self.root, image=img_left_tk)
        label_left.photo = img_left_tk
        label_left.pack(side="left", anchor="s")

        label_right = tk.Label(self.root, image=img_right_tk)
        label_right.photo = img_right_tk
        label_right.pack(side="right", anchor="s")

    def apply_filter(self, filter_type, filter_entry):
        filter_value = filter_entry.get()

        if not filter_value:
            messagebox.showerror('Hata', "Filtre değeri boş olamaz.")
            return

        filtered_students = []

        if filter_type == "yıl":
            if not filter_value.isdigit():
                messagebox.showerror('Hata', "Filtreleme yılı bir sayı olmalıdır.")
                return
            filtered_students = self.student_manager.filter_by_year(int(filter_value))
        elif filter_type == "universite":
            filtered_students = [row for row in
                                 self.student_manager.workbook.active.iter_rows(min_row=2, values_only=True) if
                                 row[1] == filter_value]
        elif filter_type == "fakulte":
            filtered_students = [row for row in
                                 self.student_manager.workbook.active.iter_rows(min_row=2, values_only=True) if
                                 row[3] == filter_value]

        if not filtered_students:
            messagebox.showinfo("Bilgi", f"Belirtilen filtre değerine ait kayıt bulunamadı.")
            return

        new_filename = f"{filter_value}_filtrelenmis.xlsx"
        self.student_manager.export_filtered_students(filtered_students, new_filename)

        messagebox.showinfo("Başarılı", f"Belirtilen filtre değerine ait kayıtlar {new_filename} dosyasına kaydedildi.")

        self.clear_frame()
        self.initialize_gui()

    def clear_frame(self):
        for widget in self.root.winfo_children():
            widget.destroy()


def main():
    # Check the license first
    if not check_license():
        time.sleep(5000)
        return  # Terminate the program if the license is not valid

    root = tk.Tk()
    script_dir = os.path.dirname(os.path.abspath(__file__))
    image_left_path = os.path.join(script_dir, "hawks.png")
    image_right_path = os.path.join(script_dir, "amal.png")

    student_manager = StudentManager()
    app = StudentManagementGUI(root, student_manager, image_left_path, image_right_path)
    root.mainloop()

if __name__ == "__main__":
    main()
